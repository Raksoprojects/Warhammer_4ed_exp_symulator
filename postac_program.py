import openpyxl

cost_table = {
    6: (25, 10),
    11: (30, 15),
    16: (40, 20),
    21: (50, 30),
    26: (70, 40),
    31: (90, 60),
    36: (120, 80),
    41: (150, 110),
    46: (190, 140),
    51: (230, 180),
    56: (280, 220),
    61: (330, 270),
    66: (390, 320),
    71: (450, 380),
    # Default value for advancements beyond 70
    float('inf'): (520, 440)
}


# Funkcja do wczytania danych z pliku Excel
def load_data(file_name):
    wb = openpyxl.load_workbook(file_name, data_only=True)  # Dodaj argument data_only=True
    sheet = wb.active
    cechy = {}
    umiejetnosci = {}
    exp = {}
    poczatkowa = {}

    # Wczytanie danych cech
    for col in range(2, 10 + 1):
        cecha = sheet.cell(row=11, column=col).value
        poczatkowa = sheet.cell(row=12, column=col).value
        rozwinieta = sheet.cell(row=13, column=col).value
        wartosc = poczatkowa + rozwinieta
        cechy[cecha] = {'poczatkowa': poczatkowa, 'rozwinieta': rozwinieta, 'wartosc': wartosc}

    # Wczytanie danych umiejętności
    a = 0     
    while a < 3:  
        for row in range(18, 31):
            umiejetnosc = sheet.cell(row=row, column=1+(a*5)).value
            if umiejetnosc is None:
                break
            cecha = sheet.cell(row=row, column=2+(a*5)).value
            wartosc = sheet.cell(row=row, column=3+(a*5)).value
            rozwinieta = sheet.cell(row=row, column=4+(a*5)).value
            suma = wartosc + rozwinieta
            umiejetnosci[umiejetnosc] = {'cecha': cecha, 'wartosc': wartosc, 'rozwinieta': rozwinieta, 'suma': suma}
        a+=1        

    # Wczytanie danych dotyczących doświadczenia
    exp['aktualne'] = sheet['P11'].value
    exp['wydane'] = sheet['Q11'].value
    exp['suma'] = sheet['R11'].value

    wb.close()
    return cechy, umiejetnosci, exp


# Funkcja do zapisu danych do pliku Excel
def save_data(file_name, cechy, umiejetnosci, exp):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.active

    # Zapis danych cech
    for col in range(2, 10 + 1):
        cecha = sheet.cell(row=11, column=col).value
        sheet.cell(row=12, column=col, value=cechy[cecha]['poczatkowa'])
        sheet.cell(row=14, column=col, value=cechy[cecha]['wartosc'])
        sheet.cell(row=13, column=col, value=cechy[cecha]['rozwinieta'])

    # Zapis danych umiejętności
    a = 0     
    while a < 3:     
        for row in range(18, 31):
            umiejetnosc = sheet.cell(row=row, column=1 + (a * 5)).value
            if umiejetnosc is not None:
                if umiejetnosci.get(umiejetnosc) is not None:
                    sheet.cell(row=row, column=2 + (a * 5), value=umiejetnosci[umiejetnosc].get('cecha', 'Default Value'))
                    sheet.cell(row=row, column=3 + (a * 5), value=umiejetnosci[umiejetnosc]['wartosc'])
                    sheet.cell(row=row, column=4 + (a * 5), value=umiejetnosci[umiejetnosc]['rozwinieta'])
                    sheet.cell(row=row, column=5 + (a * 5), value=umiejetnosci[umiejetnosc]['suma'])
        a += 1


    # Zapis danych dotyczących doświadczenia
    sheet['P11'] = exp['aktualne']
    sheet['Q11'] = exp['wydane']
    sheet['R11'] = exp['suma']

    wb.save(file_name)
    wb.close()


# Funkcja do obliczenia kosztu rozwinięcia
def calculate_total_experience(advancement_type, current_advancements, desired_advancements):
    total_experience = 0
    remaining_advancements = desired_advancements
    current_threshold = current_advancements

    while remaining_advancements > 0:
        for threshold, (char_exp, skill_exp) in cost_table.items():
            if current_threshold < threshold:
                advancements_to_threshold = min(remaining_advancements, threshold - current_threshold)
                if advancement_type == 'cecha':
                    total_experience += char_exp * advancements_to_threshold
                elif advancement_type == 'umiejetnosc':
                    total_experience += skill_exp * advancements_to_threshold
                remaining_advancements -= advancements_to_threshold
                current_threshold += advancements_to_threshold
                if remaining_advancements == 0:
                    break
    return total_experience


def wyswietl_cechy_umiejetnosci(cechy, umiejetnosci, exp):
    print("CECHY:")
    print("Cecha | Wartość | Rozwinięcia | Maks. rozwinięcia")
    for cecha, wartosci in cechy.items():
        max_rozwinięć_cechy = 0
        exp_do_wykorzystania = exp['aktualne']
        posiadane_rozwinięcia = wartosci['rozwinieta']
        while True:
            koszt_cechy = calculate_total_experience('cecha', posiadane_rozwinięcia, max_rozwinięć_cechy + 1)
            if exp_do_wykorzystania >= koszt_cechy:
                max_rozwinięć_cechy += 1
                #exp_do_wykorzystania -= koszt_cechy
            else:
                break
        print(f"{cecha.ljust(5)} | {str(wartosci['wartosc']).ljust(7)} | {str(wartosci['rozwinieta']).ljust(11)} | {max_rozwinięć_cechy}")

    print("\nUMIEJĘTNOŚCI:")
    print("Umiejętność \t\t| Wartość | Rozwinięcia | Maks. rozwinięcia")
    for umiejetnosc, dane in umiejetnosci.items():
        max_rozwinięć_umiejetności = 0
        exp_do_wykorzystania = exp['aktualne']
        posiadane_rozwinięcia = dane['rozwinieta']
        while True:
            koszt_umiejetnosci = calculate_total_experience('umiejetnosc', posiadane_rozwinięcia, max_rozwinięć_umiejetności + 1)
            if exp_do_wykorzystania >= koszt_umiejetnosci:
                max_rozwinięć_umiejetności += 1
                #exp_do_wykorzystania -= koszt_umiejetnosci
            else:
                break
        print(f"{umiejetnosc.ljust(23)} | {str(dane['wartosc']).ljust(7)} | {str(dane['rozwinieta']).ljust(11)} | {max_rozwinięć_umiejetności}")

    print("\nMaksymalne rozwinięcia uwzględniają posiadane rozwinięcia oraz cenę poszczególnych rozwinięć.")




def menu_glowne(cechy, umiejetnosci, exp):
    while True:
        print("\nMENU GŁÓWNE")
        print("------------")
        print("1. Wyświetl dane")
        print("2. Zakup rozwinięć")
        print("3. Dodaj doświadczenie")
        print("4. Wyświetl cechy, umiejętności i maksymalne rozwinięcia")
        print("Q. Wyjdź")

        wybor = input("Wybierz opcję: ").upper()

        if wybor == '1':
            print("\nCECHY:")
            print("Cecha | Początkowa | Rozwinięta | Wartość")
            for cecha, wartosci in cechy.items():
                print(f"{cecha.ljust(5)} | {str(wartosci['poczatkowa']).ljust(10)} | {str(wartosci['rozwinieta']).ljust(10)} | {wartosci['wartosc']}")

            print("\nUMIEJĘTNOŚCI:")
            print("Umiejętność \t\t| Cecha | Wartość | Rozwinięcia | Suma")
            for umiejetnosc, dane in umiejetnosci.items():
                print(f"{umiejetnosc.ljust(23)} | {dane['cecha'].ljust(5)} | {str(dane['wartosc']).ljust(7)} | {str(dane['rozwinieta']).ljust(11)} | {dane['suma']}")

            print("\nDOŚWIADCZENIE:")
            print(f"Aktualne: {exp['aktualne']}, Wydane: {exp['wydane']}, Suma: {exp['suma']}")

        elif wybor == '2':
            cecha_lub_umiejetnosc = input("Podaj nazwę cechy lub umiejętności: ")

            if cecha_lub_umiejetnosc in cechy:
                rozwinieta = cechy[cecha_lub_umiejetnosc]['rozwinieta']
                koszt = calculate_total_experience('cecha', rozwinieta, 1)
            elif cecha_lub_umiejetnosc in umiejetnosci:
                rozwinieta = umiejetnosci[cecha_lub_umiejetnosc]['rozwinieta']
                koszt = calculate_total_experience('umiejetnosc', rozwinieta, 1)
            else:
                print("Nieprawidłowa nazwa cechy lub umiejętności.")
                continue

            print(f"Koszt 1 rozwinięcia: {koszt} doświadczenia.")

            if exp['aktualne'] >= koszt:
                ilosc_rozw = int(input("Podaj ilość rozwinięć: "))
                koszt = calculate_total_experience('cecha' if cecha_lub_umiejetnosc in cechy else 'umiejetnosc', rozwinieta, ilosc_rozw)

                print(f"Koszt {ilosc_rozw} rozwinięć: {koszt} doświadczenia.    Aktualne doświadczenie: {exp['aktualne']}")
                czy_zapisac = input(f"Czy zapisać zmiany? (Wydać PD) Y/N: ")

                if czy_zapisac.upper() == 'Y':

                    if exp['aktualne'] >= koszt:
                        exp['wydane'] += koszt
                        exp['aktualne'] -= koszt
                        print("Zakupiono rozwinięcia.")
                        if cecha_lub_umiejetnosc in cechy:
                            cechy[cecha_lub_umiejetnosc]['rozwinieta'] += ilosc_rozw
                            cechy[cecha_lub_umiejetnosc]['wartosc'] += ilosc_rozw
                            for umiejetnosc, dane in umiejetnosci.items():
                                if dane['cecha'].strip("()") == cecha_lub_umiejetnosc:
                                    umiejetnosci[umiejetnosc]['wartosc'] += ilosc_rozw
                                    umiejetnosci[umiejetnosc]['suma'] += ilosc_rozw

                        elif cecha_lub_umiejetnosc in umiejetnosci:
                            umiejetnosci[cecha_lub_umiejetnosc]['rozwinieta'] += ilosc_rozw
                            umiejetnosci[cecha_lub_umiejetnosc]['suma'] += ilosc_rozw
                    else:
                        print("Brak wystarczającej ilości doświadczenia.")

                else:
                    print("Nie zapisano zmian.")
                    continue        
            else:
                print("Brak wystarczającej ilości doświadczenia.")

        elif wybor == '3':
            exp_increase = int(input("Podaj ilość dodawanego doświadczenia: "))
            exp['aktualne'] += exp_increase
            exp['suma'] = exp['aktualne'] + exp['wydane']
            print(f"Dodano {exp_increase} doświadczenia. Aktualne dostępne doświadczenie: {exp['aktualne']}. Aktualna suma doświadczenia: {exp['suma']}")

        elif wybor == '4':
            # Wyświetl cechy, umiejętności i maksymalne rozwinięcia
            wyswietl_cechy_umiejetnosci(cechy, umiejetnosci, exp)

        elif wybor == 'Q':
            save = input("Czy chcesz zapisać zmiany w danych umiejętności do pliku Excel? (T/N): ").upper()
            if save == 'T':
                save_data("karta_postaci.xlsx", cechy, umiejetnosci, exp)
                print("Zapisano zmiany w pliku Excel.")
            print("Do widzenia!")
            break

        else:
            print("Nieprawidłowy wybór. Spróbuj ponownie.")


# Główna funkcja programu
def main():
    cechy, umiejetnosci, exp = load_data("karta_postaci.xlsx")
    menu_glowne(cechy, umiejetnosci, exp)


if __name__ == "__main__":
    main()